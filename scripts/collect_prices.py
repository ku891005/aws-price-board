#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AWS 단가 수집기 v2 — RDS / EC2 (기본: 서울 ap-northeast-2)

기존 사내 스크립트(collect_rds_sqlserver_pricing.py)의 도메인 규칙을 그대로 계승합니다.
  · BYOM 에디션 병합    : databaseEdition 'Enterprise-BYOM' → 'Enterprise' + 라이선스 BYOM
  · 언번들 세대 처리     : m7i/r7i/m8i/r8i/m8a/r8a 는 라이선스 요금이 분리 공표
                          → Compute(라이선스 제외) = LI − License
  · 로컬 캐시 재사용     : _pricelist_<svc>_<region>.json (--refresh 로 재다운로드)
  · 월 환산             : 시간당 × 730
추가된 부분
  · RI 매트릭스          : 1yr/3yr × No Upfront / Partial Upfront / All Upfront (실효 시간단가)
  · EC2 동시 수집
  · 게시판용 JSON + 기존 형식 엑셀(--xlsx) 동시 출력

표준 라이브러리만으로 동작합니다(requests 불필요). --xlsx 에만 openpyxl 필요.

사용 예
  python scripts/collect_prices.py                        # 서울, JSON 출력
  python scripts/collect_prices.py --refresh              # 캐시 무시하고 재다운로드
  python scripts/collect_prices.py --xlsx                 # 엑셀도 함께 저장
  python scripts/collect_prices.py --region ap-northeast-1
  python scripts/collect_prices.py --cache-dir ./_cache --services rds
"""
from __future__ import annotations

import argparse
import collections
import datetime
import gzip
import json
import os
import ssl
import sys
import traceback
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT_DIR = os.path.join(ROOT, "docs", "data")
LOG_PATH = os.path.join(ROOT, "collect_log.txt")
_logf = None

BULK = "https://pricing.us-east-1.amazonaws.com/offers/v1.0/aws/{svc}/current/{region}/index.json"
SVC = {"rds": "AmazonRDS", "ec2": "AmazonEC2"}
REGION_NAMES = {
    "ap-northeast-2": "서울", "ap-northeast-1": "도쿄", "us-east-1": "버지니아 북부",
    "us-west-2": "오레곤", "ap-southeast-1": "싱가포르", "eu-west-1": "아일랜드",
}
HOURS = 730.0

# 라이선스 요금을 인스턴스 요금과 분리해 공표하는 세대 (기존 스크립트 계승)
UNBUNDLED = ("m7i", "r7i", "m8i", "r8i", "m8a", "r8a")

# licenseModel 원문 → 표시 라벨
LICENSE_LABELS = {
    "License included": "License Included",
    "Bring your own license": "BYOL",
    "Bring your own media": "BYOM",
    "No license required": "No License Required",
    "NA": "N/A",
    "": "N/A",
}

# RI 조합 → 데이터 키
RI_KEYS = {
    ("1yr", "No Upfront"): "ri1nu", ("1yr", "Partial Upfront"): "ri1pu",
    ("1yr", "All Upfront"): "ri1au", ("3yr", "No Upfront"): "ri3nu",
    ("3yr", "Partial Upfront"): "ri3pu", ("3yr", "All Upfront"): "ri3au",
}
YEARS = {"1yr": 1, "3yr": 3}


# ────────────────────────────── 로그 ──────────────────────────────
def log(msg=""):
    global _logf
    text = str(msg)
    try:
        print(text, flush=True)
    except Exception:
        pass
    try:
        if _logf is None:
            _logf = open(LOG_PATH, "w", encoding="utf-8")
        _logf.write(text + "\n")
        _logf.flush()
    except Exception:
        pass


# ──────────────────────────── 다운로드 ────────────────────────────
def fetch(svc_key, region, cache_dir, refresh=False):
    """Bulk JSON 다운로드 (표준 urllib + 진행률 + 로컬 캐시)."""
    svc = SVC[svc_key]
    os.makedirs(cache_dir, exist_ok=True)
    cache = os.path.join(cache_dir, f"_pricelist_{svc_key}_{region}.json")
    if os.path.isfile(cache) and not refresh:
        log(f"[캐시] 재사용: {os.path.basename(cache)} ({os.path.getsize(cache)/1048576:.1f} MB)"
            "  (--refresh 로 재다운로드)")
        with open(cache, encoding="utf-8") as f:
            return json.load(f)

    url = BULK.format(svc=svc, region=region)
    log(f"[다운로드] {url}")
    log("  파일이 큽니다(수십~수백 MB). 1~5분 걸릴 수 있습니다...")
    ctx = ssl.create_default_context()
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (aws-price-board collector)",
        "Accept-Encoding": "gzip",
    })
    buf = bytearray()
    with urllib.request.urlopen(req, timeout=900, context=ctx) as r:
        gz = r.headers.get("Content-Encoding") == "gzip"
        total = int(r.headers.get("Content-Length") or 0)
        while True:
            chunk = r.read(1024 * 512)
            if not chunk:
                break
            buf.extend(chunk)
            mb = len(buf) / 1048576
            tail = f" / {total/1048576:.1f} MB" if total else ""
            print(f"\r  받는 중... {mb:7.1f} MB{tail}", end="", flush=True)
    print()
    raw = gzip.decompress(bytes(buf)) if gz else bytes(buf)
    log(f"[완료] {len(raw)/1048576:.1f} MB")
    with open(cache, "wb") as f:
        f.write(raw)
    log(f"[캐시 저장] {os.path.basename(cache)}")
    return json.loads(raw.decode("utf-8"))


# ──────────────────────────── 공통 파싱 ────────────────────────────
def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def norm_license(v):
    v = (v or "").strip()
    return LICENSE_LABELS.get(v, v or "N/A")


def split_byom(edition):
    """'Enterprise-BYOM' → ('Enterprise', True). 기존 스크립트의 BYOM 병합 규칙."""
    ed = (edition or "").strip()
    if ed.endswith("-BYOM"):
        return ed[:-5], True
    return ed, False


def is_unbundled(instance_type):
    parts = instance_type.split(".")
    fam = parts[1] if len(parts) >= 3 else (parts[0] if parts else "")
    return fam in UNBUNDLED


def is_license_dim(attrs, desc):
    """언번들 세대의 '라이선스 요금' SKU 판별.

    기존 스크립트는 description 에 'license' 가 있으면 라이선스분으로 봤으나,
    '(License Included)' 문구까지 걸려 오탐이 발생했다.
    여기서는 언번들 세대 + 라이선스 전용 표현만 인정한다.
    """
    it = attrs.get("instanceType") or ""
    if not is_unbundled(it):
        return False
    ut = (attrs.get("usagetype") or "").lower()
    if ":license" in ut or ut.endswith("-license") or "licensefee" in ut:
        return True
    d = desc.lower()
    return ("license fee" in d) or ("license cost" in d) or \
           ("license" in d and "included" not in d and "instance hour" not in d)


def hourly_of(dims):
    """priceDimensions → (시간당 합계, 선결제 합계)"""
    h = up = 0.0
    for d in dims.values():
        usd = _num((d.get("pricePerUnit") or {}).get("USD")) or 0.0
        unit = (d.get("unit") or "").lower()
        if unit == "quantity":
            up += usd
        elif unit in ("hrs", "hours", "hour"):
            h += usd
    return h, up


def classify(attrs, desc):
    """제품 구분(RDS / RDS Custom)과 라이선스 모델 판정.

    · engineCode 400번대 = RDS Custom (SQL Server 401~407, Oracle 410~411 등)
    · SQL Server 의 RDS Custom SKU 는 licenseModel 이 'NA' 로 비어 있어
      요금 설명문의 'AWS-provided media' / 'customer-provided media(BYOM)' 로 구분한다.
    """
    code = (attrs.get("engineCode") or "").strip()
    try:
        product = "RDS Custom" if int(code) >= 400 else "RDS"
    except ValueError:
        product = "RDS"

    lic = norm_license(attrs.get("licenseModel"))
    if lic in ("N/A", ""):
        d = desc.lower()
        if "customer-provided media" in d or "(byom)" in d or "bring your own media" in d:
            lic = "BYOM"
        elif "aws-provided" in d or "(li)" in d or "license included" in d:
            lic = "License Included"
        else:
            # RDS Custom SQL Server 는 engineCode 로 최종 판정
            #   401·402·403 = AWS 제공 미디어(LI) / 405·406·407 = 고객 제공 미디어(BYOM)
            lic = {"401": "License Included", "402": "License Included", "403": "License Included",
                   "405": "BYOM", "406": "BYOM", "407": "BYOM"}.get(code, lic)
    return product, lic


def price_matrix(terms, sku, attrs):
    """SKU 하나의 온디맨드 + RI 매트릭스. 라이선스 전용 SKU는 별도 표시."""
    out = {}
    lic_only = False
    descs = []

    for t in (terms.get("OnDemand", {}).get(sku) or {}).values():
        dims = t.get("priceDimensions") or {}
        desc = " ".join((d.get("description") or "") for d in dims.values())
        h, _ = hourly_of(dims)
        if h <= 0:
            continue
        descs.append(desc)
        if is_license_dim(attrs, desc):
            lic_only = True
            out["license_od"] = round(h, 6)
        else:
            out["od"] = round(h, 6)

    for t in (terms.get("Reserved", {}).get(sku) or {}).values():
        a = t.get("termAttributes") or {}
        if a.get("OfferingClass") != "standard":
            continue
        key = RI_KEYS.get((a.get("LeaseContractLength"), a.get("PurchaseOption")))
        if not key:
            continue
        yrs = YEARS[a["LeaseContractLength"]]
        dims = t.get("priceDimensions") or {}
        desc = " ".join((d.get("description") or "") for d in dims.values())
        h, up = hourly_of(dims)
        eff = h + up / (yrs * 365 * 24)
        if eff <= 0:
            continue
        if is_license_dim(attrs, desc):
            out["license_" + key] = round(eff, 6)
        else:
            prev = out.get(key)
            if prev is None or eff < prev:
                out[key] = round(eff, 6)
    return out, lic_only, " ".join(descs)


# ──────────────────────────── RDS ────────────────────────────
def build_rds(data):
    products, terms = data.get("products", {}), data.get("terms", {})
    log(f"[RDS] 전체 SKU {len(products):,}개")

    inst = {}          # (engine, edition, license, deployment, itype) → row
    lic_fees = {}      # (itype, deployment) → 라이선스 전용 단가
    storage = []
    skipped = collections.Counter()

    for sku, p in products.items():
        a = p.get("attributes", {})
        fam = p.get("productFamily") or ""

        if fam == "Database Instance":
            it = (a.get("instanceType") or "").strip()
            if not it.startswith("db."):
                skipped["instanceType 없음"] += 1
                continue
            eng = (a.get("databaseEngine") or "").strip()
            ed, byom = split_byom(a.get("databaseEdition"))
            dep = (a.get("deploymentOption") or "").strip()

            pm, lic_only, desc = price_matrix(terms, sku, a)
            product, lic = classify(a, desc)
            if byom:
                lic = "BYOM"
            if lic_only:
                key = (it, dep)
                cur = lic_fees.get(key, {})
                for k, v in pm.items():
                    if k.startswith("license_"):
                        cur[k[len("license_"):]] = v
                lic_fees[key] = cur
                continue
            if "od" not in pm:
                skipped["온디맨드 없음"] += 1
                continue

            key = (product, eng, ed or "-", lic, dep, it)
            row = inst.get(key)
            if row is None:
                row = {
                    "sku": sku, "instanceType": it,
                    "family": it.split(".")[1] if it.count(".") >= 2 else "",
                    "product": product,
                    "engine": eng, "edition": ed or "-", "license": lic, "deployment": dep,
                    "vcpu": int(_num(a.get("vcpu")) or 0),
                    "memory": (a.get("memory") or "").strip(),
                    "processor": (a.get("physicalProcessor") or "").strip(),
                    "network": (a.get("networkPerformance") or "").strip(),
                    "usagetype": (a.get("usagetype") or "").strip(),
                    "unbundled": is_unbundled(it),
                }
                inst[key] = row
            for k, v in pm.items():
                if row.get(k) is None or v < row[k]:
                    row[k] = v

        elif fam in ("Database Storage", "Provisioned IOPS", "Provisioned Throughput",
                     "Storage Snapshot", "System Operation", "Performance Insights",
                     "RDSProxy", "ServerlessV2", "Serverless"):
            ed, byom = split_byom(a.get("databaseEdition"))
            for t in (terms.get("OnDemand", {}).get(sku) or {}).values():
                for d in (t.get("priceDimensions") or {}).values():
                    usd = _num((d.get("pricePerUnit") or {}).get("USD"))
                    if not usd:
                        continue
                    storage.append({
                        "sku": sku, "family": fam,
                        "type": (a.get("volumeType") or a.get("storageMedia")
                                 or a.get("group") or "-"),
                        "engine": (a.get("databaseEngine") or "-"),
                        "edition": ed or "-",
                        "deployment": (a.get("deploymentOption") or "-"),
                        "unit": d.get("unit", ""), "price": usd,
                        "desc": d.get("description", ""),
                    })

    # ① 라이선스 요금이 별도 SKU로 공표된 경우: Compute = LI − License
    filled = 0
    for row in inst.values():
        fee = lic_fees.get((row["instanceType"], row["deployment"]))
        if not fee or row["license"] != "License Included":
            continue
        row["licenseFee"] = fee.get("od")
        for k in ("od",) + tuple(RI_KEYS.values()):
            base, lf = row.get(k), fee.get(k)
            if base is not None and lf is not None and base > lf:
                row["c_" + k] = round(base - lf, 6)
                filled += 1

    # ② 라이선스 요금 별도 청구 판별 (실측 기반)
    #    언번들 세대(m7i·r7i·m8i·r8i·m8a·r8a)는 LI 단가와 BYOM 단가가 동일하다.
    #    = 공표 단가에 라이선스료가 포함되어 있지 않고, 라이선스는 별도 청구되며
    #      가격표에는 공표되지 않는다는 뜻. 이 경우 LI 단가가 곧 Compute 단가다.
    by_spec = collections.defaultdict(dict)
    for row in inst.values():
        by_spec[(row["product"], row["engine"], row["edition"], row["deployment"],
                 row["instanceType"])][row["license"]] = row
    sep = 0
    for group in by_spec.values():
        li, byom = group.get("License Included"), group.get("BYOM")
        if not li or not byom:
            continue
        a, b = li.get("od"), byom.get("od")
        if a is not None and b is not None and abs(a - b) < 1e-9:
            for r in (li, byom):
                r["licSep"] = True
                if r.get("c_od") is None:
                    r["c_od"] = r["od"]      # 공표 단가 = 라이선스 제외 컴퓨팅 단가
            sep += 1

    log(f"[RDS] 인스턴스 {len(inst):,}행 · 스토리지 {len(storage):,}행")
    log(f"[RDS] 라이선스 별도 SKU {len(lic_fees):,}건 · (LI−License) 산출 {filled:,}셀 · "
        f"라이선스 별도청구 스펙 {sep:,}건")
    if skipped:
        log(f"[RDS] 제외: {dict(skipped)}")
    return list(inst.values()), storage


# ──────────────────────────── EC2 ────────────────────────────
def build_ec2(data):
    products, terms = data.get("products", {}), data.get("terms", {})
    log(f"[EC2] 전체 SKU {len(products):,}개")
    rows = {}
    for sku, p in products.items():
        a = p.get("attributes", {})
        if p.get("productFamily") not in ("Compute Instance", "Compute Instance (bare metal)"):
            continue
        if a.get("capacitystatus") != "Used":
            continue
        if a.get("tenancy") not in ("Shared", "Dedicated"):
            continue
        pm, _, _desc = price_matrix(terms, sku, a)
        if "od" not in pm:
            continue
        it = (a.get("instanceType") or "").strip()
        key = (a.get("operatingSystem"), norm_license(a.get("licenseModel")),
               a.get("preInstalledSw"), a.get("tenancy"), it)
        row = rows.get(key)
        if row is None:
            row = {
                "sku": sku, "instanceType": it, "family": it.split(".")[0] if "." in it else it,
                "os": a.get("operatingSystem", ""), "license": norm_license(a.get("licenseModel")),
                "preSw": a.get("preInstalledSw", "NA"), "tenancy": a.get("tenancy", ""),
                "vcpu": int(_num(a.get("vcpu")) or 0), "memory": (a.get("memory") or "").strip(),
                "processor": (a.get("physicalProcessor") or "").strip(),
                "network": (a.get("networkPerformance") or "").strip(),
                "usagetype": (a.get("usagetype") or "").strip(),
            }
            rows[key] = row
        for k, v in pm.items():
            if row.get(k) is None or v < row[k]:
                row[k] = v
    log(f"[EC2] 인스턴스 {len(rows):,}행")
    return list(rows.values())


# ──────────────────────────── 출력 ────────────────────────────
def write_json(name, obj):
    os.makedirs(OUT_DIR, exist_ok=True)
    path = os.path.join(OUT_DIR, name)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, separators=(",", ":"))
    n = len(obj) if isinstance(obj, list) else "-"
    log(f"   docs/data/{name}  {n}행  {os.path.getsize(path)/1048576:.2f} MB")


SQL_HDR = ["구분", "스펙명", "vCPU", "Memory", "언번들세대", "배포옵션",
           "OD LI 시간당($)", "OD LI 월($)",
           "★OD BYOM 시간당($)", "★OD BYOM 월($)",
           "★RI No-Up 1yr BYOM 시간당($)", "★RI No-Up 1yr BYOM 월($)",
           "OD Compute(라이선스제외) 시간당($)", "OD Compute 월($)", "OD License분 시간당($)",
           "RI No-Up 1yr LI 시간당($)", "RI No-Up 1yr LI 월($)",
           "RI No-Up 1yr BYOL 시간당($)", "RI No-Up 1yr BYOL 월($)",
           "RI License분 시간당($)", "KEY"]
OPN_HDR = ["엔진", "에디션", "스펙명", "vCPU", "Memory", "배포옵션",
           "OD 시간당($)", "OD 월($)", "RI No-Up 1yr 시간당($)", "RI No-Up 1yr 월($)", "KEY"]

SQL_ENG = {"Enterprise": "SQLServer Enterprise", "Standard": "SQLServer Standard",
           "Web": "SQLServer Web", "Express": "SQLServer Express"}


def build_xlsx_records(rds_rows, deployment="Single-AZ"):
    """기존 엑셀 시트 형식(SQLServer / OpenSource) 레코드 생성."""
    mo = lambda v: round(v * HOURS, 2) if v else None
    idx = collections.defaultdict(dict)
    for r in rds_rows:
        if r["deployment"] != deployment or r.get("product") != "RDS":
            continue
        idx[(r["engine"], r["edition"], r["instanceType"])][r["license"]] = r

    sql, opn = [], []
    for (eng, ed, it), by_lic in sorted(idx.items()):
        li = by_lic.get("License Included")
        byom = by_lic.get("BYOM")
        byol = by_lic.get("BYOL")
        base = li or byom or byol or next(iter(by_lic.values()))
        if eng == "SQL Server":
            label = SQL_ENG.get(ed)
            if not label:
                continue
            od_li = li.get("od") if li else None
            ri_li = li.get("ri1nu") if li else None
            od_byom = byom.get("od") if byom else None
            ri_byom = byom.get("ri1nu") if byom else None
            od_comp = (li or {}).get("c_od")
            ri_comp = (li or {}).get("c_ri1nu") or ((byol or {}).get("ri1nu"))
            lic_fee = (li or {}).get("licenseFee")
            sql.append([label, it, base.get("vcpu"), base.get("memory"),
                        "Y" if base.get("unbundled") else "N", deployment,
                        od_li, mo(od_li), od_byom, mo(od_byom), ri_byom, mo(ri_byom),
                        od_comp, mo(od_comp), lic_fee,
                        ri_li, mo(ri_li), ri_comp, mo(ri_comp),
                        (li or {}).get("license_ri1nu"), f"{label}|{it}"])
        else:
            od = base.get("od")
            ri = base.get("ri1nu")
            opn.append([eng, ed, it, base.get("vcpu"), base.get("memory"), deployment,
                        od, mo(od), ri, mo(ri), f"{eng}|{ed}|{it}"])
    return sql, opn


def save_xlsx(rds_rows, region, stamp):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("[경고] openpyxl 미설치 → 엑셀 생략 (pip install openpyxl)")
        return []

    sql, opn = build_xlsx_records(rds_rows)
    NAVY, F = "1F3864", "맑은 고딕"
    HF = PatternFill("solid", fgColor=NAVY)
    thin = Side(style="thin", color="BFBFBF")
    BD = Border(thin, thin, thin, thin)
    wb = openpyxl.Workbook()

    def sheet(ws, hdr, data, widths):
        for j, h in enumerate(hdr, 1):
            c = ws.cell(1, j, h)
            c.font = Font(name=F, bold=True, color="FFFFFF")
            c.fill = HF
            c.alignment = Alignment("center", "center", wrap_text=True)
            c.border = BD
        for i, row in enumerate(data, 2):
            for j, v in enumerate(row, 1):
                c = ws.cell(i, j, v)
                c.font = Font(name=F, size=9)
                c.border = BD
                c.alignment = Alignment("left" if j in (1, 2, len(hdr)) else "center", "center")
                if isinstance(v, float):
                    c.number_format = "#,##0.0000" if v < 100 else "#,##0.00"
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A2"
        if data:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{len(data)+1}"

    ws = wb.active
    ws.title = "SQLServer"
    sheet(ws, SQL_HDR, sql,
          [22, 20, 7, 12, 10, 12, 15, 14, 17, 15, 21, 19, 22, 16, 18, 18, 16, 18, 18, 17, 30])
    sheet(wb.create_sheet("OpenSource"), OPN_HDR, opn,
          [16, 14, 20, 7, 12, 12, 15, 14, 18, 16, 26])

    ws3 = wb.create_sheet("README")
    notes = [
        ("수집 정보", ""),
        ("리전", f"{REGION_NAMES.get(region, region)} ({region})"),
        ("수집 시각", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("출처", "AWS Price List Bulk API (AmazonRDS current)"),
        ("배포 옵션", "Single-AZ 기준 (Multi-AZ는 통상 2배)"),
        ("월 환산", f"시간당 × {HOURS:g}시간"), ("", ""),
        ("컬럼 설명", ""),
        ("OD LI", "On-Demand · License Included(라이선스 포함) 요금"),
        ("★OD BYOM", "On-Demand · BYOM(Bring Your Own Media, 라이선스 미포함) 인스턴스 요금. "
                     "databaseEdition '*-BYOM' 또는 licenseModel='Bring your own media' SKU에서 수집."),
        ("★RI No-Up 1yr BYOM", "BYOM 기준 RI(No Upfront 1년). AWS 미출시 스펙은 공란."),
        ("OD Compute(라이선스제외)",
         "언번들 세대는 AWS가 License 요금을 별도 SKU로 공표 → (LI − License) = 컴퓨팅+OS 상당액. "
         "번들 세대(m5·m6i·r5·r6i·t3 등)는 분리 요금이 없어 공란이 정상."),
        ("RI No-Up 1yr BYOL", "RI(No Upfront 1년) 기준 라이선스 제외 상당액."),
        ("언번들세대", "Y = " + " / ".join(UNBUNDLED)),
        ("", ""),
        ("주의", "번들 세대는 AWS가 라이선스 비용을 인스턴스 요금에 합쳐 단일 요금으로만 제공하므로 "
                "라이선스 제외 단가가 원천적으로 공표되지 않습니다(공란이 정상)."),
    ]
    for i, (a, b) in enumerate(notes, 1):
        ca = ws3.cell(i, 1, a)
        ca.font = Font(name=F, bold=(b == ""), color=(NAVY if b == "" else "000000"))
        cb = ws3.cell(i, 2, b)
        cb.font = Font(name=F, size=10)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 95

    out = os.path.join(ROOT, f"RDS_SQLServer_Pricing_{region}_{stamp}.xlsx")
    wb.save(out)
    log(f"   {out}  (SQLServer {len(sql)}행 · OpenSource {len(opn)}행)")
    return [out]


# ──────────────────────────── main ────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="AWS RDS/EC2 단가 수집기 v2")
    ap.add_argument("--region", default=os.environ.get("AWS_PRICE_REGION", "ap-northeast-2"))
    ap.add_argument("--services", default="rds,ec2", help="rds,ec2 중 선택 (콤마 구분)")
    ap.add_argument("--refresh", action="store_true", help="캐시 무시하고 재다운로드")
    ap.add_argument("--cache-dir", default=os.path.join(ROOT, "_cache"))
    ap.add_argument("--xlsx", action="store_true", help="기존 형식 엑셀도 저장")
    args = ap.parse_args()

    region = args.region
    svcs = [s.strip() for s in args.services.split(",") if s.strip()]
    log("=" * 70)
    log("AWS 단가 수집기 v2  (게시판 JSON + 기존 엑셀 형식)")
    log(f"  파이썬 {sys.version.split()[0]}")
    log(f"  ★ 리전 {REGION_NAMES.get(region, region)} ({region})")
    log(f"  대상  {', '.join(svcs)}")
    log(f"  캐시  {args.cache_dir}")
    log("=" * 70)

    rds_rows, rds_storage, ec2_rows = [], [], []
    if "rds" in svcs:
        rds_rows, rds_storage = build_rds(fetch("rds", region, args.cache_dir, args.refresh))
    if "ec2" in svcs:
        ec2_rows = build_ec2(fetch("ec2", region, args.cache_dir, args.refresh))

    rds_rows.sort(key=lambda r: (r["engine"], r["edition"], r["license"], r["vcpu"], r["od"]))
    ec2_rows.sort(key=lambda r: (r["os"], r["license"], r["vcpu"], r["od"]))

    log("[저장]")
    if "rds" in svcs:
        write_json("rds.json", rds_rows)
        write_json("rds_storage.json", rds_storage)
    if "ec2" in svcs:
        write_json("ec2.json", ec2_rows)
    sampled = [s for s in ("rds", "ec2") if s not in svcs
               and os.path.isfile(os.path.join(OUT_DIR, f"{s}.json"))]
    if sampled:
        log(f"[주의] 이번 실행에서 수집하지 않은 데이터셋: {', '.join(sampled)} "
            "(기존 파일 유지 — 샘플일 수 있음)")
    write_json("meta.json", {
        "region": region, "regionName": REGION_NAMES.get(region, region),
        "notCollected": sampled,
        "updatedAt": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
        "hoursPerMonth": HOURS, "unbundled": list(UNBUNDLED),
        "counts": {"rds": len(rds_rows), "ec2": len(ec2_rows), "rdsStorage": len(rds_storage)},
        "source": "AWS Price List Bulk API (public)",
    })

    if args.xlsx and rds_rows:
        save_xlsx(rds_rows, region, datetime.datetime.now().strftime("%Y%m%d"))

    # 채움 현황 (기존 스크립트의 검증 출력 계승)
    if rds_rows:
        sqls = [r for r in rds_rows if r["engine"] == "SQL Server" and r["deployment"] == "Single-AZ"]
        by = lambda lic, k: sum(1 for r in sqls if r["license"] == lic and r.get(k) is not None)
        log("")
        log("[검증] SQL Server · Single-AZ 채움 현황")
        log(f"   행 수                 : {len(sqls)}")
        log(f"   OD LI                 : {by('License Included','od')}")
        log(f"   OD BYOM (★)           : {by('BYOM','od')}")
        log(f"   RI No-Up 1yr LI       : {by('License Included','ri1nu')}")
        log(f"   RI No-Up 1yr BYOM (★) : {by('BYOM','ri1nu')}   (AWS 미출시 시 0)")
        log(f"   OD Compute(라이선스제외): {sum(1 for r in sqls if r.get('c_od') is not None)}")
        log(f"   라이선스 별도청구 표시  : {sum(1 for r in sqls if r.get('licSep'))}")
        fams = sorted({r['family'] for r in sqls if r['license'] == 'BYOM'})
        if fams:
            log(f"   BYOM 제공 패밀리        : {', '.join('db.'+f for f in fams)}")
    log("")
    log("[완료]")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        log("\n[오류 발생] 아래 내용을 그대로 알려주세요:")
        log(traceback.format_exc())
    finally:
        try:
            if _logf:
                _logf.close()
        except Exception:
            pass
        if sys.stdin and sys.stdin.isatty():
            try:
                input("\n엔터를 누르면 창이 닫힙니다...")
            except Exception:
                pass
